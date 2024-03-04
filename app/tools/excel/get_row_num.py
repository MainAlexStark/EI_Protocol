from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

from loguru import logger

from datetime import datetime, timedelta

from ..word.Word import Word

WORD = Word()

def get_row_num(path: str, date_to_find: str):
    # Загрузка существующего файла
        workbook = load_workbook(filename=path)

        #Выбор активного листа
        worksheet = workbook.active

        rows = []

        for i in range(1, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=i, column=10).value
            if cell_value == date_to_find:
                rows.append(i)

        workbook.close()

        return rows