from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

from loguru import logger

from datetime import datetime, timedelta





def add_var_to_excel(args):
        logger.debug('start')
        
        logger.debug(args)
        
        num_protocol = args['num_protocol'].split('-',1)[0] + '-' + args['work_place'].split('-')[0].strip() + '-' + args['num_protocol'].split('-',1)[1]

        # Исходная строка с датой
        date_str = args['inspection_date']

        # Преобразуем строку в объект datetime
        date_obj = datetime.strptime(date_str, "%d.%m.%Y")

        # Вычитаем один день
        new_date = date_obj - timedelta(days=1)

        # Вычитаем один год
        new_date = new_date.replace(year=new_date.year - int(args['interval']))

        # Преобразуем объект datetime обратно в строку
        new_date_str = new_date.strftime("%d.%m.%Y")

        if args['unfit'] == False:
                unfit = 'Пригодно'
        else:
                unfit = 'Непригодно'

        # Загрузка существующего файла
        workbook = load_workbook(filename=args['path_to_excel'])

        #Выбор активного листа
        worksheet = workbook.active

        # Находим первую пустую строку
        empty_row = worksheet.max_row + 1

        # Заполняем ячейки новыми значениями
        worksheet['A' + str(empty_row)] = num_protocol
        worksheet['B' + str(empty_row)] = 1
        worksheet['C' + str(empty_row)] = args['scale'].rsplit(' ',1)[1]
        worksheet['F' + str(empty_row)] = args['scale'].rsplit(' ',1)[0]
        worksheet['H' + str(empty_row)] = args['num_scale']
        worksheet['J' + str(empty_row)] = args['inspection_date']
        worksheet['K' + str(empty_row)] = new_date_str
        worksheet['M' + str(empty_row)] = args['method']
        worksheet['N' + str(empty_row)] = unfit
        worksheet['O' + str(empty_row)] = args['temperature'] + ' °C'
        worksheet['P' + str(empty_row)] = args['pressure'] + ' кПа'
        worksheet['Q' + str(empty_row)] = args['humidity'] + ' %'
        worksheet['R' + str(empty_row)] = args['change_temperature']
        worksheet['V' + str(empty_row)] = args['standarts_briefly']
        worksheet['AG' + str(empty_row)] = args['company']
        worksheet['AI' + str(empty_row)] = '1'
        worksheet['AJ' + str(empty_row)] = args['verificationer']
        worksheet['AS' + str(empty_row)] = args['INN']
        worksheet['AT' + str(empty_row)] = args['legal_address']
        worksheet['AU' + str(empty_row)] = args['inspection_address']
        
        
        words = ['A','B','C','F','H','J','K','M','N','Q','P','O','R','V','AG','AI','AJ','AS','AT','AU','AV']

        # Настраиваем шрифт и выравнивание

        font = Font(name='Times New Roman', size=8, bold=False, italic=False, color='000000')
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        

        # Присвяиваем шрифт и выравнивание клеткам
        for word in words:
                worksheet[word + str(empty_row)].font = font
                worksheet[word + str(empty_row)].alignment = alignment

        # Сохраняем изменения в файле Excel
        workbook.save(filename=args['path_to_excel'])

