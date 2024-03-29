from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

import os

from loguru import logger

from datetime import datetime, timedelta

import shutil

def extract_value(text, start_text, end_text):
    try:
        if text.find(start_text) != -1:
            start = text.find(start_text) + len(start_text)

            end = text.find(end_text)

            if end_text == '\n':
                value = text[start:]
            else:
                value = text[start:end]

            return value.strip()
        
    except Exception as e:
        logger.error(f'Ошибка при получении значения: {e}')

class Excel():
    
    def open_document(self, path: str):
        
        try:

            # Загрузка существующего файла
            workbook = load_workbook(filename=path)
            
            return workbook
        
        except Exception as e:
            logger.error(f'Ошибка при открытии шаблона')

    def create_template(self, path: str):
        logger.debug('Создание шаблона протокола Excel')

        try:

            # Путь к исходному файлу
            source_file = path
            # Путь к папке, в которую нужно скопировать файл
            destination_folder = "app/templates/Excel"

            # Название файла
            file_name = os.path.basename(source_file).rsplit('.',1)[0]
            scale = file_name.split(' ',1)[1].split('№')[0]
            FIF = file_name.rsplit('(',1)[1].split(')')[0]

            # Копируем файл в целевую папку
            shutil.copy(source_file, destination_folder)

            # Получаем полный путь к скопированному файлу
            new_file_path = os.path.join(destination_folder, os.path.basename(source_file))

            # Новое название файла
            new_file_name = f"{scale} {FIF}.xlsx"

            # Полный путь к файлу с новым названием
            new_file_path_with_name = os.path.join(destination_folder, new_file_name)

            logger.debug(f'\n{new_file_path_with_name}')

            # Переименовываем файл
            os.rename(new_file_path, new_file_path_with_name)

            logger.success('Шаблон Excel создан')

            return True
        
        except Exception as e:
            logger.error('Ошибка при создании Excel шаблона')
            return e

    def make_new_protocol(self, args):
        logger.debug("Создание Excel протокола")

        path_to_excel_file = f'app\\templates\\Excel\\{args['scale']}.xlsx'
        
        workbook = self.open_document(path=path_to_excel_file)
        
        if workbook is not None:

            try:

                #Выбор активного листа
                worksheet = workbook['Данные']

                worksheet['B1'] = args['scale']
                worksheet['B2'] = args['num_protocol']
                worksheet['B4'] = args['num_scale']
                worksheet['B5'] = args['readings']
                worksheet['B6'] = args['company']
                worksheet['B7'] = args['inspection_address']
                worksheet['B8'] = args['work_place_combo'].split('-')[0]
                worksheet['B9'] = args['inspection_date']
                worksheet['B10'] = args['temperature']
                worksheet['B11'] = args['humidity']
                worksheet['B12'] = args['pressure']
                worksheet['B13'] = args['temperature_liquid']
                worksheet['B14'] = args['temperature_liquid']
                
                worksheet['C8'] = args['verificationer_combo']
                
                worksheet['D8'] = args['standarts']

                scale = args['scale'].rsplit(' ',1)[0]
                fif = args['scale'].rsplit(' ',1)[1]

                full = f'{args['save_path']}/{args['num_protocol']} {scale} №{args['num_scale']} ({fif}).xlsx'

                # Сохраняем изменения в файле Excel
                workbook.save(filename=full)

                workbook.close()

                logger.success("Протокол Excel создан")
                
                return args
            except Exception as e:
                logger.error(f'Ошибка при создании Excel протокола:{e}')
        else:
            logger.error('Неправильный путь к файлу')

    def add_args_to_excel_journal(self, args):
        logger.debug('start')
        
        logger.debug(args)
        
        num_protocol = args['num_protocol'].split('-',1)[0] + '-' + args['work_place_combo'].split('-')[0].strip() + '-' + args['num_protocol'].split('-',1)[1]

        # Исходная строка с датой
        date_str = args['inspection_date']

        # Преобразуем строку в объект datetime
        date_obj = datetime.strptime(date_str, "%d.%m.%Y")

        # Вычитаем один день
        new_date = date_obj - timedelta(days=1)

        # Вычитаем один год
        new_date = new_date.replace(year=new_date.year - int(args['interval']))

        # Преобразуем объект datetime обратно в строку
        new_date_str = new_date.strftime("%d.%m.%Y")

        if args['unfit_button'] == False:
                unfit = 'Пригодно'
        else:
                unfit = 'Непригодно'

        # Загрузка существующего файла
        workbook = self.open_document(path=args['path_to_excel_jounal'])
        
        if workbook is not None:

            #Выбор активного листа
            worksheet = workbook.active

            # Находим первую пустую строку
            empty_row = worksheet.max_row + 1

            # Заполняем ячейки новыми значениями
            worksheet['A' + str(empty_row)] = num_protocol
            worksheet['B' + str(empty_row)] = 1
            worksheet['C' + str(empty_row)] = args['scale'].rsplit(' ',1)[1]
            worksheet['F' + str(empty_row)] = args['scale'].rsplit(' ',1)[0]
            worksheet['H' + str(empty_row)] = args['num_scale']
            worksheet['J' + str(empty_row)] = args['inspection_date']
            worksheet['K' + str(empty_row)] = new_date_str
            worksheet['M' + str(empty_row)] = args['method']
            worksheet['N' + str(empty_row)] = unfit
            worksheet['O' + str(empty_row)] = args['temperature'] + ' °C'
            worksheet['P' + str(empty_row)] = args['pressure'] + ' кПа'
            worksheet['Q' + str(empty_row)] = args['humidity'] + ' %'
            worksheet['R' + str(empty_row)] = args['change_temperature']
            worksheet['V' + str(empty_row)] = args['standarts_briefly']
            worksheet['AG' + str(empty_row)] = args['company']
            worksheet['AI' + str(empty_row)] = '1'
            worksheet['AJ' + str(empty_row)] = args['verificationer_combo']
            worksheet['AS' + str(empty_row)] = args['INN']
            worksheet['AT' + str(empty_row)] = args['legal_address']
            worksheet['AU' + str(empty_row)] = args['inspection_address']
            
            
            words = ['A','B','C','F','H','J','K','M','N','Q','P','O','R','V','AG','AI','AJ','AS','AT','AU','AV']

            # Настраиваем шрифт и выравнивание

            font = Font(name='Times New Roman', size=8, bold=False, italic=False, color='000000')
            alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            

            # Присвяиваем шрифт и выравнивание клеткам
            for word in words:
                    worksheet[word + str(empty_row)].font = font
                    worksheet[word + str(empty_row)].alignment = alignment

            # Сохраняем изменения в файле Excel
            workbook.save(filename=args['path_to_excel'])
        
        
    def get_date_rows_from_excel_journal(self, path: str, date_to_find: str):
        # Загрузка существующего файла
        workbook = self.open_document(path=path)
        
        if workbook is not None:

            #Выбор активного листа
            worksheet = workbook.active

            rows = []

            for i in range(1, worksheet.max_row + 1):
                cell_value = worksheet.cell(row=i, column=10).value
                if cell_value == date_to_find:
                    rows.append(i)

            workbook.close()

            return rows
    
    
    def get_args_from_excel(self, path: str, row: str):

        try:

            # Загрузка существующего файла
            workbook = load_workbook(filename=path)
            
            if workbook is not None:

                #Выбор активного листа
                worksheet = workbook.active

                args = {}

                # Заполняем ячейки новыми значениями
                args['num_protocol'] = worksheet['A' + row].value
                args['FIF'] = worksheet['C' + row].value
                args['scale'] = worksheet['F' + row].value
                args['num_scale'] = worksheet['H' + row].value
                args['inspection_date'] = worksheet['J' + row].value
                args['temperature'] = worksheet['O' + row].value.replace('°C','')
                args['pressure'] = worksheet['P' + row].value.replace('кПа','')
                args['humidity'] = worksheet['Q' + row].value.replace('%','')
                args['standarts'] = worksheet['V' + row].value
                args['company'] = worksheet['AG' + row].value
                args['verificationer'] = worksheet['AJ' + row].value
                args['INN'] = worksheet['AS' + row].value
                args['legal_address'] = worksheet['AT' + row].value
                args['inspection_address'] = worksheet['AU' + row].value

                args['scale'] += ' ' +  args['FIF']

                args['work_place'] = args['num_protocol'].split('-',1)[1].split('-',1)[0] + '-'

                flag = False
                for value in args.values():
                    if not(len(value) > 0):
                        flag = True


                args['use_excel'] = False

                if worksheet['N' + row].value == 'Непригодно':
                    args['unfit'] = True
                else:
                    args['unfit'] = False


                if flag:
                    return False
                else:
                    return args

        except Exception as e:
            logger.error(f'Ошибка при получении значений из excel файла: {e}')